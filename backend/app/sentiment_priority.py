import pandas as pd
from app.gmail_fetch import fetch_emails
from app.sentiment_priority import classify_email
from openpyxl import load_workbook
from openpyxl.styles import Font

# Step 1: Fetch emails
print("Fetching emails...")
emails = fetch_emails(n=20)
print(f"Total emails fetched: {len(emails)}")

# Step 2: Classify emails
classified_emails = []
for idx, email in enumerate(emails, start=1):
    result = classify_email(email)
    classified_emails.append(result)

    print(f"\n--- Email {idx} ---")
    print(f"Subject: {result['subject']}")
    print(f"From: {result['sender']}")
    print(f"Type: {result['type']}")
    print(f"Sentiment: {result['sentiment']}")
    print(f"Priority: {result['priority']}")

# Step 3: Sorting order
priority_order = {"Urgent": 0, "Not Urgent": 1}

# Step 4: Split into sheets by type
output_file = "emails.xlsx"
with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
    for t in ["support", "query", "help", "request", "spam"]:
        filtered = [e for e in classified_emails if e["type"] == t]
        if not filtered:
            continue

        # Sort inside each type by priority
        filtered.sort(key=lambda x: priority_order.get(x["priority"], 2))
        df = pd.DataFrame(filtered)

        # Write to sheet
        df.to_excel(writer, sheet_name=t.capitalize(), index=False)

# Step 5: Formatting headers (bold + autofilter)
wb = load_workbook(output_file)
for sheet in wb.sheetnames:
    ws = wb[sheet]

    # Bold headers
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # Add autofilter
    ws.auto_filter.ref = ws.dimensions

wb.save(output_file)

print(f"\n✅ Emails classified, sorted into sheets, and saved to {output_file}")
